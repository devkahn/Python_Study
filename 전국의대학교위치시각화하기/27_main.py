#27 전국의 대학교 위치 시각화하기

## pip install folium - 지도를 시각화할 수 있는 라이브러리
## pip install openpyxl - 엑셀을 사용하기 위한 라이브러리



### 전국의 대학교 주소록 엑셀 파일 자료 받기
# kess.kedi.re.kr/index  --> 자료실


### 판다스에서 학교명과 주소 찾는 코드 만들기
import pandas as pd

fileFolder = r'pjt27_전국의대학교위치시각화하기\res'
filePath = fileFolder +'\\고등교육기관 하반기 주소록(2021).xlsx'

df_from_excel = pd.read_excel(filePath, engine='openpyxl')

df_from_excel.columns = df_from_excel.loc[4].tolist()

df_from_excel = df_from_excel.drop(index=list(range(0,5)))

print(df_from_excel.head())
print(df_from_excel['학교명'].values)
print(df_from_excel['주소'].values)



### 오픈API를 이용해 주소를 좌표로 변환하는 코드 만들기

# https://www.vworld.kr/dev/v4dv_geocoderguide2_s001.do
# 회원가입 후 인증키 발급



### 주소를 좌표로 변환하는 코드
import requests

url = 'http://api.vworld.kr/req/address?'
params = 'service=address&request=getcoord&version=2.0&crs=epsg:4326&address=%ED%9A%A8%EB%A0%B9%EB%A1%9C72%EA%B8%B8%2060&refine=true&simple=false&format=json&type='
road_type ='ROAD' #도로명 주소
road_type2 = 'PARCEL' #지번 주소
address = '&address='
keys = '&key='
primary_key = 'A3BAF90A-5678-3AA0-9E78-5A54354FA2AB' # 발급받은 인증키를 붙여넣기

def request_geo(road):
    x=0
    y=0
    
    
    reqURL = url+params+road_type+address+road+keys+primary_key
    print(reqURL)
    
    page = requests.get(url+params+road_type+address+road+keys+primary_key)
    json_data = page.json()
    if json_data['response']['status'] == 'OK':
        x = json_data['response']['result']['point']['x']
        y = json_data['response']['result']['point']['y']
        
    return x, y

x,y = request_geo('경기도 시흥시 산기대학로 237 (정왕동, 한국산업기술대학교)')

print(f'x값: {x}')
print(f'y값: {y}')


### 엑셀에서 읽은 학교명과 주소를 x,y좌표로 만들고 엑셀 파일 생성
from openpyxl import load_workbook
from openpyxl import Workbook
import re

try:
    wb = load_workbook(filePath, data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active


university_list = df_from_excel['학교명'].to_list()
address_list = df_from_excel['주소'].to_list()

# 주소에서() 괄호 부분을 삭제
# API를 활용하여 주소를 좌표로 변환합니다.
# 학교명, 주소XY의 순서대로 엑셀에 저장합니다.


### 특정 학교의 위치에 마커를 표시하는 코드 만들기
import folium

map = folium.Map(location=[37,127], zoom_start=7)
marker = folium.Marker([37.341435483, 126.733026596],
                       popup='한국공학대학교',
                       icon=folium.Icon(color='blue'))

marker.add_to(map)
map.save(r'pjt27_전국의대학교위치시각화하기\res\uni_map.html')
