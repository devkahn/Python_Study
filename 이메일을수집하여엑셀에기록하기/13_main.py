# 이메일을 수집하여 엑셀에 기록하기

##<<<정규 표현식>>>##
## . : 하나의 문자와 일치
## [ ] :[]안의 문자열은 순서와 상관없이 포함된 문자와 일지
## [^] : []안에 포함되지 않는 문자 중 하나를 의미한다.
## ^ : 문자열의 시작 위치를 나타냅니다. 여러 줄에서 처리하는 때에는 각 줄의 시작을 나타낸다.
## $ : 문자열의 마지막 위치 또는 개행문자의 바로 앞 위치를 나타낸다.
## ( ) : 괄호 안의 일치되는 부분을 묶어서 사용할 수 있습니다.
## \1 : !~9까지의 숫자를 표현합니다,.
## \\ : 영어 소문자, 언더바, 영어 대문자, 숫자를 표현합니다.
## * : 바로 앞의 패턴이 0번 이상 일치 합니다.
## [1,2] :  바로 앞의 패턴이 최소 1번 최대 2번 일치 합니다.
## ? : 바로 앞의 패턴의 0 또는 1번 일치 합니다.
## + : 바로 앞의 패턴이 번 이상 일치 합니다.
## | : 앞의 패턴 또는 뒤의 패턴 중 하나와 일치 합니다.

### 이메일 형식을 추출하는 코드 만들기
import re

test_string = """
aaa@bbb.com
123@abc.co.kr
test@hello.ke
ok@ok.co.kr
ok@ok.com
ok@ok.co.kr
ok@ok.co.kr
no.co.kr
no.kr
"""

results = re.findall(r'[\w\.-]+@[\w\.-]+', test_string)
print(results)


### 리스트에서 중복 내용 제거하는 코드 만들기
results = list(set(results))
print(results)


### 사이트에서 이메일 수집하는 코드 만들기
import requests

url = 'https://www.hani.co.kr/arti/politics/politics_general/1076813.html?_ns=t1'
headers = {
    'User-Agent' : 'Mozilla/5.0',
    'Content-Type' : 'text/html; charset=utf-8'
}

response = requests.get(url, headers=headers)
results = re.findall(r'[\w\.-]+@[\w\.-]+[.]+[\w\.-]', response.text)
results = list(set(results))

print(results)


### 수집한 이메일 주소를 엑셀에 저장하는 코드 만들기
from openpyxl import load_workbook
from openpyxl import Workbook

saveFolderPath = r'04_AutomationProgram\pjt13_이메일을수집하여엑셀에기록하기\result'

try:
    wb = load_workbook(saveFolderPath +'\\email.xlsx', data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active
    
for result in results:
    sheet.append([result])
    
wb.save(saveFolderPath +'\\email.xlsx')
